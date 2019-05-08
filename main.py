#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import csv
import pandas as pd
from pysqlite3 import dbapi2 as sqlite3

from os import path
from pandas import read_csv
from pandas.io.excel import ExcelWriter
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Protection
from openpyxl.worksheet.protection import SheetProtection
from copy import copy
from flask import Flask, request, jsonify

import constants
import card_reader
import database

global CURRENT_PATH
CURRENT_PATH = path.dirname(path.realpath(__file__))

def create_app():
    app = Flask(__name__)

    database.init_app(app)

    return app

app = create_app()

# define user's card id
# member_card_list = {
#     "01120312E418BA0E": "Duc",
#     "0101011459180432": "Thao"
# }

@app.route('/read_card', methods=['GET'])
def read_card():
    if request.method != 'GET':
        return {}

    response_json = {}

    records = card_reader.read_card(CURRENT_PATH)
    records_len = len(records)

    if not records:
        response_json.update({
            'status': 'error',
            'content': 'No card or can\'t read card.'
        })

    db = database.get_db()
    cur = db.cursor()
    for i, r in enumerate(records):
        if (i == records_len-1):
            cur.execute(constants.PRAGMA_KEY)
            cur.execute('SELECT * FROM record WHERE id = ?', (r.get('id'),))
            result = cur.fetchone()
            if result is not None:
                r['charge'] = dict(result).get('charge')
        tb_insert = '''INSERT OR IGNORE INTO record VALUES
            (:id, :date, :in_line, :in_sta, :out_line, :out_sta, :charge,
            :balance, :process, :memo, :is_saved, :created_at)'''
        cur.execute(constants.PRAGMA_KEY)
        cur.execute(tb_insert, r)
        db.commit()
    
    response_json.update({
        'status': 'success',
        'content': records
    })

    return jsonify(response_json)

@app.route('/get_history', methods=['GET'])
def get_history():
    response_json = {
        'status': 'error',
        'content': 'Error or No saved data.'
    }

    if request.method != 'GET':
        return response_json

    data = []

    db = database.get_db()
    cur = db.cursor()
    tb_get = 'SELECT * FROM record WHERE is_saved'
    cur.execute(constants.PRAGMA_KEY)
    cur.execute(tb_get)
    rows = cur.fetchall()
    data = [dict(r) for r in rows]

    if not data:
        response_json.update({
            'status': 'error',
            'content': 'No saved history.'
        })
    
    response_json.update({
        'status': 'success',
        'content': data
    })

    return jsonify(response_json)

@app.route('/save_records', methods=['POST'])
def save_records():
    response_json = {
        'status': 'error',
        'content': 'Error! Cannot save data.'
    }

    if request.method != 'POST':
        return response_json
    records = request.get_json()['records']
    
    db = database.get_db()
    cur = db.cursor()
    tb_save = 'UPDATE record SET is_saved = ?, memo = ? WHERE id = ?'
    for r in records:
        cur.execute(constants.PRAGMA_KEY)
        cur.execute(tb_save, (True, r.get('memo', ''), r.get('id')))
        db.commit()

    response_json.update({
        'status': 'success',
        'content': 'OK'
    })
    return jsonify(response_json)

@app.route('/delete_record/<string:id>', methods=['DELETE'])
def delete_record(id):
    response_json = {
        'status': 'error',
        'content': 'Error! Can\'t delete record or record doesn\'t exist.'
    }

    if request.method != 'DELETE':
        return response_json

    db = database.get_db()
    cur = db.cursor()
    tb_delete = 'UPDATE record SET is_saved = ? WHERE id = ?'
    cur.execute(constants.PRAGMA_KEY)
    cur.execute(tb_delete, (False, id))
    db.commit()

    response_json.update({
        'status': 'success',
        'content': id
    })
    return jsonify(response_json)

@app.route('/export_records', methods=['POST'])
def export_records():
    response_json = {
        'status': 'error',
        'content': 'Error! Can\'t export records to Excel Transport Record file.'
    }

    if request.method != 'POST':
        return response_json
    export_directory = request.get_json()['export_directory']
    records = request.get_json()['records']
    
    # get user name from card id
    # with open(input_file_path, 'r', encoding=ENCODING_JA) as csv_file:
    #     card_id_raw = [r for r in csv.reader(csv_file)][0][0]
    #     card_id = card_id_raw.partition('ID=')[2]
    # excelFileName = member_card_list.get(card_id, 'MemberNotFound')

    # df_in = read_csv(input_file_path, skiprows=1, encoding=constants.ENCODING_JA)

    # change data in each cell by using pandas.read_excel (not working)
    # df_out = read_excel(output_file_path,
    #     sheet_name=SHEET_NAME, skiprows=4,
    #     usecols="B:E", header=None)
    # df_out.loc[0, 1] = '10'

    # create new Excel file by using xlsxwriter engine
    # writer = ExcelWriter(output_file_path, engine='xlsxwriter')
    # df_in.to_excel(writer, sheet_name='Sheet1')
    # worksheet = writer.sheets['Sheet1']
    # worksheet.set_column('B:K', 18)
    # worksheet.set_column('A:A', 5)
    # worksheet.set_column('C:C', 8)
    # worksheet.set_column('F:F', 8)
    # writer.save()

    # create file using openpyxl
    # wb = load_workbook(output_file_path)
    # template_path dev env
    template_path = path.join(CURRENT_PATH, 'template', constants.TEMPLATE_FILE_NAME)
    # template_path prod env
    # template_path = path.join(CURRENT_PATH, 'resources', 'template', TEMPLATE_FILE_NAME)
    wb = load_workbook(template_path)
    ws = wb[constants.SHEET_NAME]
    ws.protection = SheetProtection(sheet=True, insertRows=False, password=constants.PASSWORD)

    cell_protection = Protection(locked=False)
    ws.cell(row=1, column=4).protection = cell_protection
    ws.cell(row=1, column=5).protection = cell_protection
    ws.cell(row=2, column=5).protection = cell_protection

    i = 0
    for row in records:
        # if int(row[5]) <= 0:
        #     continue
    #     insert 1 row and format
    #     ws.insert_rows(5, 1)
    #     for c in list(map(chr, range(ord('A'), ord('L')+1))):
    #         from_cell = ws[c+'6']
    #         to_cell = ws[c+'5']
    #         to_cell._style = copy(from_cell._style)
    #         to_cell.value = from_cell.value

        ws.cell(row=5+i, column=2).value = row['date']
        description = "バス" if not row['in_sta'] and not row['out_sta'] and "バス" in row['process'] \
            else "{0} から {1} まで".format(row['in_sta'], row['out_sta'])
        if row['memo']:
            description += "\n" + row['memo']
        ws.cell(row=5+i, column=3).value = description
        ws.cell(row=5+i, column=3).alignment = Alignment(wrap_text=True)
        ws.cell(row=5+i, column=3).protection = cell_protection
        ws.cell(row=5+i, column=5).value = row['charge']
        i += 1

    if i < 32:
        for r in range(5+i, 32):
            for c in range(2, 12):
                ws.cell(row=r, column=c).protection = cell_protection

    # wb.save(output_file_path)
    wb.save(path.join(export_directory, constants.TEMPLATE_FILE_NAME))

    response_json.update({
        'status': 'success',
        'content': 'Export successful.'
    })
    return jsonify(response_json)

if __name__ == '__main__':
    app.run()
